"""
Unit tests for the GH-6 Master Data XLSX writer helpers.

Tests the pure-Python module-level helpers in TRACT_Geolocation_Formatter:
- _MASTER_DATA_MAPPING (constant)
- _master_data_output_path
- _read_country_list_from_template
- _write_master_data_xlsx

Requires openpyxl. The plugin's vendored copy bootstraps via __init__.py
(GH-6 Task 1) — importing tract_geolocation_formatter first puts vendor/ on
sys.path, then openpyxl resolves. The main module additionally requires
qgis.PyQt to be importable. Tests skip gracefully if any import fails
(same pattern as test_self_intersection_coordinates.py and
test_build_expression_value.py).
"""

import hashlib
import os
import tempfile
import unittest

try:
    # Importing the plugin package first triggers __init__.py's vendor
    # bootstrap so openpyxl resolves from the vendored copy.
    import tract_geolocation_formatter  # noqa: F401
    import openpyxl
    from tract_geolocation_formatter.TRACT_Geolocation_Formatter import (
        _MASTER_DATA_MAPPING,
        _master_data_output_path,
        _read_country_list_from_template,
        _write_master_data_xlsx,
    )

    HELPERS_AVAILABLE = True
except ImportError:
    HELPERS_AVAILABLE = False


# Bundled template paths (post-Task 1 location)
_PLUGIN_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "tract_geolocation_formatter",
)
_TEMPLATES_DIR = os.path.join(_PLUGIN_DIR, "templates")
_FARMS_TEMPLATE = os.path.join(_TEMPLATES_DIR, "farms_master_data_template.xlsx")
_FARMER_GROUPS_TEMPLATE = os.path.join(_TEMPLATES_DIR, "farmer_group_master_data_template.xlsx")


def _sha256(path):
    """Compute SHA-256 of a file for integrity comparison."""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


@unittest.skipUnless(HELPERS_AVAILABLE, "Requires openpyxl (vendored via Task 1)")
class TestMasterDataOutputPath(unittest.TestCase):
    """Pure string manipulation — no filesystem touch."""

    def test_geojson_extension(self):
        self.assertEqual(
            _master_data_output_path("/tmp/mali_nodes.geojson", "farms"),
            "/tmp/mali_nodes_master_data_farms.xlsx",
        )

    def test_json_extension(self):
        self.assertEqual(
            _master_data_output_path("/tmp/mali_nodes.json", "farmer_groups"),
            "/tmp/mali_nodes_master_data_farmer_groups.xlsx",
        )

    def test_path_with_no_extension(self):
        # os.path.splitext returns (path, "") for a path without an extension.
        self.assertEqual(
            _master_data_output_path("/tmp/data", "farms"),
            "/tmp/data_master_data_farms.xlsx",
        )


@unittest.skipUnless(HELPERS_AVAILABLE, "Requires openpyxl (vendored via Task 1)")
class TestMasterDataCellMapping(unittest.TestCase):
    """Confirm _MASTER_DATA_MAPPING matches the actual template column layout."""

    def test_farms_mapping_matches_template(self):
        mapping = _MASTER_DATA_MAPPING["farms"]
        self.assertEqual(mapping["sheet_name"], "2.Farms_Template")
        self.assertEqual(mapping["name_col"], "B")
        self.assertEqual(mapping["country_col"], "E")
        self.assertEqual(mapping["ref_id_col"], "I")

        wb = openpyxl.load_workbook(_FARMS_TEMPLATE, read_only=True, data_only=True)
        try:
            self.assertIn(mapping["sheet_name"], wb.sheetnames)
            ws = wb[mapping["sheet_name"]]
            row2 = {
                cell.column_letter: cell.value
                for cell in next(ws.iter_rows(min_row=2, max_row=2))
            }
            self.assertEqual(row2[mapping["name_col"]], "Farm Name")
            self.assertEqual(row2[mapping["country_col"]], "Country")
            self.assertEqual(row2[mapping["ref_id_col"]], "Reference_ID")
        finally:
            wb.close()

    def test_farmer_groups_mapping_matches_template(self):
        mapping = _MASTER_DATA_MAPPING["farmer_groups"]
        self.assertEqual(mapping["sheet_name"], "2. Farmer_Groups Template")
        self.assertEqual(mapping["name_col"], "A")
        self.assertEqual(mapping["country_col"], "C")
        self.assertEqual(mapping["ref_id_col"], "G")

        wb = openpyxl.load_workbook(_FARMER_GROUPS_TEMPLATE, read_only=True, data_only=True)
        try:
            self.assertIn(mapping["sheet_name"], wb.sheetnames)
            ws = wb[mapping["sheet_name"]]
            row2 = {
                cell.column_letter: cell.value
                for cell in next(ws.iter_rows(min_row=2, max_row=2))
            }
            self.assertEqual(row2[mapping["name_col"]], "Farmer_Group_Name")
            self.assertEqual(row2[mapping["country_col"]], "Country")
            self.assertEqual(row2[mapping["ref_id_col"]], "Reference_ID")
        finally:
            wb.close()


@unittest.skipUnless(HELPERS_AVAILABLE, "Requires openpyxl (vendored via Task 1)")
class TestReadCountryList(unittest.TestCase):
    """AC-5 unit-level: 250 countries in TRACT order, first Afghanistan, last Zimbabwe."""

    def test_read_country_list_from_farms_template(self):
        countries = _read_country_list_from_template(_FARMS_TEMPLATE)
        self.assertEqual(len(countries), 250)
        self.assertEqual(countries[0], "Afghanistan")
        self.assertEqual(countries[-1], "Zimbabwe")

    def test_read_country_list_identical_across_templates(self):
        """Both bundled templates carry the same country list."""
        farms_list = _read_country_list_from_template(_FARMS_TEMPLATE)
        fg_list = _read_country_list_from_template(_FARMER_GROUPS_TEMPLATE)
        self.assertEqual(farms_list, fg_list)


@unittest.skipUnless(HELPERS_AVAILABLE, "Requires openpyxl (vendored via Task 1)")
class TestWriteMasterDataXlsx(unittest.TestCase):
    """Integration tests for the writer — read back the produced file with openpyxl."""

    def _write_and_open(self, master_data_type, country, unique_node_ids):
        """Write to a temp path, open with openpyxl, return (path, workbook).

        Caller is responsible for closing the workbook and deleting the path.
        """
        template = _FARMS_TEMPLATE if master_data_type == "farms" else _FARMER_GROUPS_TEMPLATE
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name
        try:
            _write_master_data_xlsx(template, output_path, master_data_type, country, unique_node_ids)
            wb = openpyxl.load_workbook(output_path, data_only=True)
            return output_path, wb
        except Exception:
            os.unlink(output_path)
            raise

    def test_write_farms_single_node(self):
        """AC-2: single unique NodeID → one data row at row 3 in B/E/I."""
        path, wb = self._write_and_open("farms", "Brazil", ["NODE_001"])
        try:
            ws = wb["2.Farms_Template"]
            self.assertEqual(ws["B3"].value, "NODE_001")
            self.assertEqual(ws["E3"].value, "Brazil")
            self.assertEqual(ws["I3"].value, "NODE_001")
            # Row 4 entirely empty (only 1 row written)
            self.assertIsNone(ws["B4"].value)
            self.assertIsNone(ws["E4"].value)
            self.assertIsNone(ws["I4"].value)
        finally:
            wb.close()
            os.unlink(path)

    def test_write_farms_multiple_nodes_order_preserved(self):
        """AC-6: caller passes already-deduped list; writer preserves order."""
        path, wb = self._write_and_open("farms", "Colombia", ["A", "B", "C"])
        try:
            ws = wb["2.Farms_Template"]
            self.assertEqual(ws["B3"].value, "A")
            self.assertEqual(ws["B4"].value, "B")
            self.assertEqual(ws["B5"].value, "C")
            # Country same on all 3 rows; Reference_ID equals Farm Name on every row
            for r in (3, 4, 5):
                self.assertEqual(ws[f"E{r}"].value, "Colombia")
                self.assertEqual(ws[f"I{r}"].value, ws[f"B{r}"].value)
            # Row 6 empty
            self.assertIsNone(ws["B6"].value)
        finally:
            wb.close()
            os.unlink(path)

    def test_write_farmer_groups_cell_positions(self):
        """AC-3: farmer-groups uses columns A/C/G, not B/E/I."""
        path, wb = self._write_and_open("farmer_groups", "Côte d'Ivoire", ["GRP_1", "GRP_2"])
        try:
            ws = wb["2. Farmer_Groups Template"]
            self.assertEqual(ws["A3"].value, "GRP_1")
            self.assertEqual(ws["C3"].value, "Côte d'Ivoire")
            self.assertEqual(ws["G3"].value, "GRP_1")
            self.assertEqual(ws["A4"].value, "GRP_2")
            self.assertEqual(ws["C4"].value, "Côte d'Ivoire")
            self.assertEqual(ws["G4"].value, "GRP_2")
            # B/E/I are NOT the farmer-groups columns — must be empty
            self.assertIsNone(ws["B3"].value)
            self.assertIsNone(ws["E3"].value)
            self.assertIsNone(ws["I3"].value)
        finally:
            wb.close()
            os.unlink(path)

    def test_write_preserves_all_sheets(self):
        """AC-4: all 7 sheets present in the output, names byte-identical to the template."""
        path, wb = self._write_and_open("farms", "Brazil", ["NODE"])
        try:
            expected_sheets = [
                "0. Introduction",
                "1. ReadMe",
                "2.Farms_Template",
                ">>>",
                "3.Farms_Sample data",
                "4. Country_List",
                "Config",
            ]
            self.assertEqual(wb.sheetnames, expected_sheets)
        finally:
            wb.close()
            os.unlink(path)

    def test_write_does_not_modify_template_on_disk(self):
        """The bundled template file is byte-identical before and after a write."""
        before = _sha256(_FARMS_TEMPLATE)
        path, wb = self._write_and_open("farms", "Brazil", ["NODE_X"])
        try:
            wb.close()
        finally:
            os.unlink(path)
        after = _sha256(_FARMS_TEMPLATE)
        self.assertEqual(before, after)


if __name__ == "__main__":
    unittest.main()
